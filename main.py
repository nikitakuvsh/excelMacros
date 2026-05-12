from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import input_file
import os
from colorama import Fore, Style, init
from summary_generator import generate_summary_sheet

init()

# =========================
# CONFIG
# =========================
FILE_PATH = input_file.input_file()
base = os.path.splitext(FILE_PATH)[0]
OUTPUT_FILE = f"{base}_generated.xlsx"

SHEET_SIP = "Расчётная модель SIP"
SHEET_PERF = "Данные по перфомансу"
SHEET_RESULT = "Расчёт"

# KPI mapping
# KPI 1 -> NSV Total
# KPI 2 -> NSV Team
# KPI 3 -> LSV/t Total
# KPI 4 -> LSV/t Team
# KPI 5 -> Earnings
# KPI 6 -> Reserved

KPI_MAPPING = {
    "NSV Total": "KPI 1",
    "NSV Team": "KPI 2",
    "LSV/t Total": "KPI 3",
    "LSV/t Team": "KPI 4",
    "Earnings": "KPI 5",
    "Nsv Prem + Total": "KPI 6",
}

QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

# =========================
# LOAD WORKBOOK
# =========================
wb = load_workbook(FILE_PATH, data_only=True)

sip_ws = wb[SHEET_SIP]
perf_ws = wb[SHEET_PERF]

# recreate result sheet
if SHEET_RESULT in wb.sheetnames:
    del wb[SHEET_RESULT]

result_ws = wb.create_sheet(SHEET_RESULT)

# =========================
# HELPERS
# =========================

def normalize_percent(value):

    if value is None:
        return None

    # если Excel formula
    if isinstance(value, str) and value.startswith("="):
        return None

    # строки
    if isinstance(value, str):

        value = value.replace("%", "")
        value = value.replace(",", ".").strip()

        try:
            value = float(value)
        except:
            return None

        # 113.46 -> 1.1346
        if value > 1:
            value = value / 100

        return value

    # числа Excel
    try:
        return float(value)
    except:
        return None
    
def normalize_earnings(value):

    if value is None:
        return None

    try:
        s = str(value).replace('%', '').replace(',', '.').strip()
        v = float(s)
    except:
        return None

    # 112 -> 1.12
    if v > 2:
        v /= 100

    return v

def find_quarter_blocks(ws):
    """
    Ищет блоки кварталов по всему листу.
    """

    blocks = {}

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):

            val = ws.cell(row, col).value

            if val is None:
                continue

            val = str(val)

            if "Q1" in val:
                blocks["Q1"] = row
            elif "Q2" in val:
                blocks["Q2"] = row
            elif "Q3" in val:
                blocks["Q3"] = row
            elif "Q4" in val:
                blocks["Q4"] = row

    return blocks


def extract_perf_data(ws):
    """
    Reads all quarters from performance sheet.

    Returns structure:

    {
        'Q1': {
            'Belarus': {
                'Perf NSV': 113.46,
                'Perf LSV/t': 100.86
            },
            'Total PN': {...}
        }
    }
    """

    quarter_blocks = find_quarter_blocks(ws)
    data = {}

    for quarter, start_row in quarter_blocks.items():

        header_row = start_row + 1
        data_start = start_row + 2

        headers = {}

        for col in range(1, ws.max_column + 1):
            h = ws.cell(header_row, col).value

            if h:
                headers[str(h).strip()] = col
            

        marketplace_col = headers.get("marketplace")
        perf_nsv_col = headers.get("Perf NSV")
        perf_lsv_col = headers.get("Perf LSV/t")
        earnings_col = headers.get("Perf Earnings")
        nsv_prem_col = headers.get("Perf Nsv Prem + Total")

        if not marketplace_col:
            continue

        quarter_data = {}

        row = data_start

        while row <= ws.max_row:

            team = ws.cell(row, marketplace_col).value

            if team is None:
                break

            team = str(team).strip()
            
            earnings_raw = None
            if earnings_col:
                earnings_raw = ws.cell(row, earnings_col).value

            quarter_data[team] = {
                "Perf NSV": normalize_percent(ws.cell(row, perf_nsv_col).value),
                "Perf LSV/t": normalize_percent(ws.cell(row, perf_lsv_col).value),
                "Perf Earnings": normalize_earnings(earnings_raw) if earnings_col else None,
                "Perf Nsv Prem + Total": normalize_earnings(ws.cell(row, nsv_prem_col).value) if nsv_prem_col else None
            }

            row += 1

        data[quarter] = quarter_data

    return data


def find_all_sip_matrices(ws):
    """
    Находит все SIP матрицы.
    """

    matrices = {}

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):

            val = ws.cell(row, col).value

            if val is None:
                continue

            val = str(val).strip()

            if not val.startswith("KPI"):
                continue

            # KPI название
            kpi_name = val

            # ищем метрику справа
            metric_name = ws.cell(row, col + 1).value
            quarter = ws.cell(row, col + 3).value

            if metric_name is None or quarter is None:
                continue

            metric_name = str(metric_name).strip()
            quarter = str(quarter).strip()

            performance_values = []
            payout_values = []

            # данные начинаются через 2 строки
            data_row = row + 2

            while True:

                perf_val = ws.cell(data_row, col + 1).value
                payout_val = ws.cell(data_row, col + 3).value

                if perf_val is None:
                    break

                if isinstance(perf_val, str):
                    break

                try:
                    perf_num = normalize_percent(perf_val)
                    payout_num = normalize_percent(payout_val)
                except:
                    break

                performance_values.append(perf_num)
                payout_values.append(payout_num)

                data_row += 1

            matrices[(kpi_name, metric_name, quarter)] = {
                "performance_values": performance_values,
                "payout_values": payout_values,
            }

    return matrices


def get_payout(matrix, performance):

    # если performance отсутствует
    if performance is None:
        return 0

    perf_values = matrix.get("performance_values", [])
    payout_values = matrix.get("payout_values", [])

    best_payout = 0

    for perf_threshold, payout in zip(perf_values, payout_values):

        if perf_threshold is None:
            continue

        if payout is None:
            continue

        try:
            if float(performance) >= float(perf_threshold):
                best_payout = payout
        except:
            continue

    return best_payout


# =========================
# READ DATA
# =========================
performance_data = extract_perf_data(perf_ws)
matrices = find_all_sip_matrices(sip_ws)

# teams list from Q1
teams = []

if "Q1" in performance_data:
    base_teams = [x for x in performance_data["Q1"].keys() if x != "Total PN"]
    teams = ["Total PN"] + base_teams

# =========================
# STYLES
# =========================
header_fill = PatternFill("solid", fgColor="D9EAF7")
bold_font = Font(bold=True)

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# =========================
# GENERATE RESULT PAGE
# =========================
current_row = 1

for team in teams:

    # -------------------------
    # HEADER
    # -------------------------

    result_ws.cell(current_row, 1).value = ""
    result_ws.cell(current_row + 1, 1).value = "KPI 1"
    result_ws.cell(current_row + 2, 1).value = "KPI 2"
    result_ws.cell(current_row + 3, 1).value = "KPI 3"
    result_ws.cell(current_row + 4, 1).value = "KPI 4"
    result_ws.cell(current_row + 5, 1).value = "KPI 5"
    result_ws.cell(current_row + 6, 1).value = "KPI 6"

    result_ws.cell(current_row, 2).value = team

    result_ws.cell(current_row + 1, 2).value = "NSV Total"
    result_ws.cell(current_row + 2, 2).value = "NSV Team"
    result_ws.cell(current_row + 3, 2).value = "LSV/t Total"
    result_ws.cell(current_row + 4, 2).value = "LSV/t Team"
    result_ws.cell(current_row + 5, 2).value = "Earnings"
    result_ws.cell(current_row + 6, 2).value = "Nsv Prem + Total"

    headers = ["Q1", "Q2", "Q3", "Q4", "FY(100%)", "Add Bonus", "FY (200%)", "Total"]

    start_col = 3

    for idx, h in enumerate(headers):
        result_ws.cell(current_row, start_col + idx).value = h

    # -------------------------
    # QUARTER CALCULATIONS
    # -------------------------

    for q_index, quarter in enumerate(QUARTERS):

        col = start_col + q_index

        if quarter not in performance_data:
            continue

        q_data = performance_data[quarter]
        total_data = q_data.get("Total PN")

        if total_data is None: continue

        if team == "Total PN":
            team_data = total_data
        else:
            if team not in q_data: continue
            team_data = q_data[team]

        team_data = q_data[team]
        total_data = q_data.get("Total PN")
        if total_data is None: continue

        # KPI 1 = NSV Total
        perf_nsv_total = total_data["Perf NSV"]

        matrix_key = ("KPI 1", "NSV Total", quarter)

        if matrix_key in matrices:
            payout = get_payout(matrices[matrix_key], perf_nsv_total)
            result_ws.cell(current_row + 1, col).value = payout

        # KPI 2 = NSV Team
        perf_nsv_team = team_data["Perf NSV"]

        matrix_key = ("KPI 2", "NSV Team", quarter)

        if matrix_key in matrices:
            payout = get_payout(matrices[matrix_key], perf_nsv_team)
            result_ws.cell(current_row + 2, col).value = payout

        # KPI 3 = LSV/t Total
        perf_lsv_total = total_data["Perf LSV/t"]

        matrix_key = ("KPI 3", "LSV/t Total", quarter)

        if matrix_key in matrices:
            payout = get_payout(matrices[matrix_key], perf_lsv_total)
            result_ws.cell(current_row + 3, col).value = payout

        # KPI 4 = LSV/t Team
        perf_lsv_team = team_data["Perf LSV/t"]

        matrix_key = ("KPI 4", "LSV/t Team", quarter)

        if matrix_key in matrices:
            payout = get_payout(matrices[matrix_key], perf_lsv_team)
            result_ws.cell(current_row + 4, col).value = payout

       # KPI 5 = Earnings
        perf_earnings = (
            team_data.get("Perf Earnings")
            if team_data.get("Perf Earnings") is not None else
            team_data.get("Earnings fcst")
            if team_data.get("Earnings fcst") is not None else
            team_data.get("Earnings plan")
        )

        matrix_key = ("KPI 5", "Earnings", quarter)

        if matrix_key in matrices:

            matrix = matrices[matrix_key]
            payout = get_payout(matrix, perf_earnings)
            result_ws.cell(current_row + 5, col).value = payout

        # KPI 6 = Nsv Prem + Total
        perf_nsv_prem = team_data.get("Perf Nsv Prem + Total")
        maxtix_key = ("KPI 6", "Nsv Prem + Total", quarter)

        if matrix_key in matrices:
            matrix = matrices[matrix_key]
            payout = get_payout(matrix, perf_nsv_prem)
            result_ws.cell(current_row + 6, col).value = payout

        # -------------------------
        # FY (200%) CALC (based on Q4 performance)
        # -------------------------

        # -------------------------
        # FY (200%) CALC (FIXED)
        # -------------------------

        if quarter == "Q4":

            fy_col = start_col + 6  # FY (200%)

            perf_map = {
                "KPI 1": total_data["Perf NSV"],
                "KPI 2": team_data["Perf NSV"],
                "KPI 3": total_data["Perf LSV/t"],
                "KPI 4": team_data["Perf LSV/t"],
                "KPI 5": total_data.get("Perf Earnings"),
                "KPI 6": total_data.get("Perf Nsv Prem + Total"),
            }

            row_map = {
                "KPI 1": current_row + 1,
                "KPI 2": current_row + 2,
                "KPI 3": current_row + 3,
                "KPI 4": current_row + 4,
                "KPI 5": current_row + 5,
                "KPI 6": current_row + 6,
            }

            def find_fy_matrix(kpi):
                for key, matrix in matrices.items():
                    if key[0] == kpi and "FY" in str(key[2]).upper():
                        return matrix
                return None

            for kpi, perf_value in perf_map.items():
                if perf_value is None: continue

                fy_matrix = find_fy_matrix(kpi)

                if fy_matrix and perf_value is not None:
                    # нормализация шкалы FY
                    fy_input = float(perf_value)

                    fy200 = get_payout(fy_matrix, fy_input)
                else:
                    fy200 = 0

                result_ws.cell(row_map[kpi], fy_col).value = fy200
    

    def safe(v): return float(v) if v is not None else 0

    kpi_rows = {
        "KPI 1": current_row + 1,
        "KPI 2": current_row + 2,
        "KPI 3": current_row + 3,
        "KPI 4": current_row + 4,
        "KPI 5": current_row + 5,
        "KPI 6": current_row + 6,
    }

    for kpi, row in kpi_rows.items():
        q1 = result_ws.cell(row, 3).value
        q2 = result_ws.cell(row, 4).value
        q3 = result_ws.cell(row, 5).value
        q4 = result_ws.cell(row, 6).value

        fy100 = (
            safe(q1) * 3 + safe(q2) * 3 + safe(q3) * 3 + safe(q4) * 4
        ) / 13
    
        result_ws.cell(row, 7).value = fy100

    for kpi, row in kpi_rows.items():

        fy100 = result_ws.cell(row, 7).value
        add_bonus = result_ws.cell(row, 8).value
        fy200 = result_ws.cell(row, 9).value

        total = safe(fy100) + safe(add_bonus) + safe(fy200)

        result_ws.cell(row, 10).value = total

    # -------------------------
    # STYLING
    # -------------------------

    yellow_fill = PatternFill("solid", fgColor="FFF200")
    blue_fill = PatternFill("solid", fgColor="D9EAF7")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")

    # заголовки кварталов
    for c in range(3, 11):
        result_ws.cell(current_row, c).fill = blue_fill
        result_ws.cell(current_row, c).font = bold_font

    # название команды
    result_ws.cell(current_row, 2).fill = yellow_fill
    result_ws.cell(current_row, 2).font = bold_font

    # KPI и метрики
    kpi_rows_range = range(current_row + 1, current_row + 7)
    for i, r in enumerate(kpi_rows_range):
        is_bold = (i % 2 == 0)
        for c in range(1, 11):
            cell = result_ws.cell(r,c)
            cell.font = bold_font if is_bold else Font(bold=False)

    # закрашиваем значения
    for r in range(current_row + 1, current_row + 7):
        for c in range(3, 11):

            cell = result_ws.cell(r, c)

            # пустые значения -> 0%
            if cell.value is None:
                cell.value = 0

            # FY (200%) и Total серые
            if c in [9, 10]:
                cell.fill = gray_fill
            else:
                cell.fill = blue_fill

            # проценты
            cell.number_format = '0.0%'

            # жирный для Total KPI
            metric_name = result_ws.cell(r, 2).value

            if metric_name in ["NSV Total", "LSV/t Total"]:
                cell.font = bold_font

            cell.border = border

    # границы вообще всей таблицы
    for r in range(current_row, current_row + 7):
        for c in range(1, 11):
            result_ws.cell(r, c).border = border

# -------------------------
# TOTAL ROW (SUM KPI 1–6)
# -------------------------

    total_row = current_row + 7  # строка ниже KPI 6

    result_ws.cell(total_row, 1).value = ""
    result_ws.cell(total_row, 1).font = bold_font

    for c in range(3, 11):  # Q1..Total
        col_sum = 0

        for r in range(current_row + 1, current_row + 7):  # KPI 1–6
            val = result_ws.cell(r, c).value

            if val is None:
                continue

            try:
                col_sum += float(val)
            except:
                continue

        cell = result_ws.cell(total_row, c)
        cell.value = col_sum
        cell.number_format = '0.0%'
        cell.font = bold_font
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.border = border
    # spacing between matrices
    current_row += 10

# =========================
# AUTO WIDTH
# =========================
for col in range(1, result_ws.max_column + 1):

    max_len = 0
    col_letter = get_column_letter(col)

    for row in range(1, result_ws.max_row + 1):
        val = result_ws.cell(row, col).value

        if val is not None:
            max_len = max(max_len, len(str(val)))

    result_ws.column_dimensions[col_letter].width = max_len + 5

# =========================
# SAVE
# =========================
try:
    generate_summary_sheet(wb, result_ws, performance_data, matrices)
    wb.save(OUTPUT_FILE)
    print(Fore.GREEN + Style.BRIGHT + "Готово!" + Style.RESET_ALL)
    print(f'Сгенерированный расчёт находится в сохраннёном файле {OUTPUT_FILE} на страничке Расчёт')
    print('Также была сгенерирована сводная таблица. Она находится на странице Сводная FY26')
    
except PermissionError:
    print(Fore.RED + Style.BRIGHT + 'Невозможно перезаписать файл. Сначала необходимо его закрыть или удалить. Расчёт не был завершён' + Style.RESET_ALL)