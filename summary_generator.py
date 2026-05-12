from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)


def generate_summary_sheet(
        wb,
        result_ws,
        performance_data,
        matrices
):

    SHEET_NAME = "Сводная FY26"

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)

    # =====================================================
    # HELPERS
    # =====================================================

    def safe(v):
        try:
            return float(v)
        except:
            return 0
        
    def get_result_value(team_name, metric_name, column_name):

        headers = {}

        # ==========================================
        # HEADERS
        # ==========================================

        for col in range(1, result_ws.max_column + 1):

            value = result_ws.cell(1, col).value

            if value:
                headers[str(value).strip()] = col

        target_col = headers.get(column_name)

        if not target_col:
            return 0

        # ==========================================
        # FIND TEAM BLOCK
        # ==========================================

        team_row = None

        for row in range(1, result_ws.max_row + 1):

            value = result_ws.cell(row, 2).value

            if value is None:
                continue

            if str(value).strip() == str(team_name).strip():

                team_row = row

                break

        if not team_row:

            return 0

        # ==========================================
        # FIND KPI INSIDE TEAM BLOCK
        # ==========================================

        for row in range(team_row + 1, team_row + 10):

            metric = result_ws.cell(row, 2).value

            if metric is None:
                continue

            if str(metric).strip() == str(metric_name).strip():

                value = result_ws.cell(row, target_col).value

                return safe(value)

        return 0

    def get_payout(matrix_key, performance):

        matrix = matrices.get(matrix_key)

        if not matrix:
            return 0

        if performance is None:
            return 0

        perf_values = matrix.get("performance_values", [])
        payout_values = matrix.get("payout_values", [])

        best_payout = 0

        for perf_threshold, payout in zip(perf_values, payout_values):

            if perf_threshold is None:
                continue

            if payout is None:
                continue

            try:
                if float(performance) >= float(perf_threshold):
                    best_payout = payout
            except:
                continue

        return best_payout

    # =====================================================
    # STYLES
    # =====================================================

    bold = Font(bold=True)
    big_bold = Font(bold=True, size=12)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    medium_border = Border(
        left=medium,
        right=medium,
        top=medium,
        bottom=medium
    )

    pink_fill = PatternFill("solid", fgColor="F4CCCC")
    gray_fill = PatternFill("solid", fgColor="E7E6E6")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    green_font = Font(
        bold=True,
        color="008000"
    )

    red_font = Font(
        bold=True,
        color="C00000"
    )

    orange_font = Font(
        bold=True,
        color="E26B0A"
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    vertical = Alignment(
        textRotation=90,
        horizontal="center",
        vertical="center"
    )

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================

    widths = {
        "A": 6,
        "B": 18,
        "C": 12,
        "D": 10,
        "E": 12,
        "F": 10,
        "G": 12,
        "H": 10,
        "I": 12,
        "J": 10,
        "K": 12,
        "L": 10,
        "M": 12,
        "N": 10,
        "O": 14,
        "P": 10,
        "Q": 12,
        "R": 10,
    }

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # =====================================================
    # MERGES
    # =====================================================

    ws.merge_cells("C1:H1")
    ws.merge_cells("I1:R1")

    ws["C1"] = "По кварталам (с учётом компенсации по итогам года)"
    ws["I1"] = "Годовая премия сверх плана"

    # =====================================================
    # HEADERS
    # =====================================================

    ws["B2"] = "FY'26"

    header_groups = [
        ("C2", "NSV"),
        ("E2", "LSV/t"),
        ("G2", "100%"),
        ("I2", "NSV"),
        ("K2", "LSV/t"),
        ("M2", "EARNINGS"),
        ("O2", "Nsv Prem + Total"),
        ("Q2", "Total"),
    ]

    for cell, text in header_groups:

        col = ws[cell].column
        row = ws[cell].row

        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + 1
        )

        ws[cell] = text

    # =====================================================
    # SUBHEADERS
    # =====================================================

    for col in range(3, 19, 2):

        ws.cell(3, col).value = "Performance"
        ws.cell(3, col + 1).value = "Payout"

    # =====================================================
    # HEADER STYLE
    # =====================================================

    for row in range(1, 4):
        for col in range(2, 19):

            cell = ws.cell(row, col)

            cell.font = bold
            cell.alignment = center
            cell.border = medium_border
    
    for cell in ["C1", "I1"]:
        ws[cell].font = Font(bold=True, size=11)
        ws[cell].fill = white_fill

    # =====================================================
    # TEAMS
    # =====================================================

    marketplace_teams = [
        "Belarus",
        "North",
        "Moscow",
        "South",
        "Center",
        "Siberia",
        "Far East",
    ]

    account_teams = [
        "X5",
        "Magnit & Dixy",
        "Lenta",
        "HMs",
        "LAR",
        "OMNI",
        "D-Com",
        "Любимчик 2.0",
    ]

    # =====================================================
    # VERTICAL LABELS
    # =====================================================

    ws.merge_cells("A4:A10")
    ws.merge_cells("A11:A19")

    ws["A4"] = "Marketplaces"
    ws["A11"] = "Account teams"

    for c in ["A4", "A11"]:
        ws[c].font = bold
        ws[c].alignment = vertical
        ws[c].fill = gray_fill
        ws[c].border = medium_border

    # =====================================================
    # DATA
    # =====================================================

    q4_data = performance_data.get("Q4", {})
    total_pn_data = q4_data.get("Total PN", {})

    all_teams = marketplace_teams + account_teams + ["Total PN"]

    # =====================================================
    # MERGED EARNINGS
    # =====================================================

    earnings_perf = safe(total_pn_data.get("Perf Earnings"))

    earnings_payout = get_payout(
        ("KPI 5", "Earnings", "FY (200%)"),
        earnings_perf
    )

    ws.merge_cells("M4:M18")
    ws.merge_cells("N4:N18")

    ws["M4"] = earnings_perf
    ws["N4"] = earnings_payout

    # =====================================================
    # MERGED KPI 6
    # =====================================================

    kpi6_perf = safe(
        total_pn_data.get("Perf Nsv Prem + Total")
    )

    kpi6_payout = get_payout(
        ("KPI 6", "Nsv Prem + Total", "FY (200%)"),
        kpi6_perf
    )

    ws.merge_cells("O4:O18")
    ws.merge_cells("P4:P18")

    ws["O4"] = kpi6_perf
    ws["P4"] = kpi6_payout

    # =====================================================
    # STYLE MERGED
    # =====================================================

    for cell_name in ["M4", "N4", "O4", "P4"]:

        cell = ws[cell_name]

        cell.font = big_bold
        cell.alignment = center
        cell.fill = gray_fill
        cell.border = border
        cell.number_format = '0.00%'

        # if col in [4, 6, 8, 10, 12, 14, 16, 18]:
        #     if value >= 0.10: cell.font = green_font
        #     elif value <= 0: cell.font = red_font
        #     else: cell.font = orange_font

        # else:
        #     if value >= 1: cell.font = green_font
        #     else: cell.font = red_font

    # =====================================================
    # WRITE TEAMS
    # =====================================================

    excel_row = 4

    for team in all_teams:

        ws.cell(excel_row, 2).value = team
        ws.cell(excel_row, 2).font = bold
        ws.cell(excel_row, 2).border = border

        team_data = 0
        total_pn_data = q4_data.get("Total PN", {})

        if team == "Total PN":
            team_data = total_pn_data
        else:
            team_data = q4_data.get(team)

        if team != "Total PN" and not team_data:
            excel_row += 1
            continue

        # =================================================
        # PERFORMANCE
        # =================================================

        nsv_perf = safe(team_data.get("Perf NSV"))
        lsv_perf = safe(team_data.get("Perf LSV/t"))

        total_pn_nsv_perf = safe(
            total_pn_data.get("Perf NSV")
        )

        total_pn_lsv_perf = safe(
            total_pn_data.get("Perf LSV/t")
        )

        # =================================================
        # Q4 PAYOUTS
        # =================================================

        nsv_payout = get_result_value(team, "NSV Team", "FY(100%)")
        lsv_payout = get_result_value(team, "LSV/t Team", "FY(100%)")

        total_pn_nsv_payout = get_payout(
            ("KPI 1", "NSV Total", "Q4"),
            total_pn_nsv_perf
        )

        total_pn_lsv_payout = get_payout(
            ("KPI 3", "LSV/t Total", "Q4"),
            total_pn_lsv_perf
        )

        # =================================================
        # 100%
        # =================================================

        payout_100 = (
            nsv_payout
            + lsv_payout
            + total_pn_nsv_payout
            + total_pn_lsv_payout
        )

        perf_100 = payout_100 / 0.30

        # =================================================
        # FY (200%)
        # =================================================

       # =================================================
# FY (200%)
# =================================================

        # =================================================
# FY (100%) payout из листа расчёта
# =================================================

        fy_nsv_payout = get_result_value(
            team,
            "NSV Team",
            "FY(100%)"
        )

        fy_lsv_payout = get_result_value(
            team,
            "LSV/t Team",
            "FY(100%)"
        )

        fy_200_nsv_payout = get_result_value(team, "NSV Team", "FY (200%)")
        fy_200_lsv_payout = get_result_value(team, "LSV/t Team", "FY (200%)")

        # =================================================
        # TOTAL
        # =================================================

        # total_payout = (
        #     payout_100
        #     + fy_nsv_payout
        #     + fy_lsv_payout
        #     + earnings_payout
        # )

        total_payout = (
            payout_100
            + fy_200_nsv_payout
            + fy_200_lsv_payout
            + earnings_payout
        )

        total_perf = total_payout / 0.30

        # =================================================
        # VALUES
        # =================================================

        values = [

            nsv_perf,
            nsv_payout,

            lsv_perf,
            lsv_payout,

            perf_100,
            payout_100,

            nsv_perf,
            fy_200_nsv_payout,

            lsv_perf,
            fy_200_lsv_payout,

            total_perf,
            total_payout,
        ]

        col = 3

        for value in values:

            while col in [13, 14, 15, 16]:
                col += 1

            cell = ws.cell(excel_row, col)

            cell.value = value
            cell.font = bold
            cell.alignment = center
            cell.border = border
            cell.number_format = '0.00%'

            if col in [4, 6, 8, 10, 12, 14, 16, 18]:
                if value >= 0.10: cell.font = green_font
                elif value <= 0: cell.font = red_font
                else: cell.font = orange_font
            
            else:
                if value >= 1: cell.font = green_font
                else: cell.font = red_font

            if col % 2 == 1:
                cell.fill = pink_fill
            else:
                cell.fill = gray_fill

            col += 1

        excel_row += 1

    # =====================================================
    # TOTAL PN
    # =====================================================

    total_row = excel_row

    TOTAL_PN_100_PAYOUT = 0.285
    TOTAL_PN_TOTAL_PAYOUT = 0.385

    total_pn_100_perf = (
        TOTAL_PN_100_PAYOUT / 0.30
    )

    total_pn_total_perf = (
        TOTAL_PN_TOTAL_PAYOUT / 0.30
    )

    total_pn_nsv_payout = get_payout(
        ("KPI 1", "NSV Total", "Q4"),
        total_pn_nsv_perf
    )

    total_pn_lsv_payout = get_payout(
        ("KPI 3", "LSV/t Total", "Q4"),
        total_pn_lsv_perf
    )

    total_pn_fy_nsv_payout = safe(
        total_pn_data.get("FY NSV Total Payout")
    )

    total_pn_fy_lsv_payout = safe(
        total_pn_data.get("FY LSV/t Total Payout")
    )

    total_pn_values = [

        total_pn_nsv_perf,
        total_pn_nsv_payout,

        total_pn_lsv_perf,
        total_pn_lsv_payout,

        total_pn_100_perf,
        TOTAL_PN_100_PAYOUT,

        total_pn_nsv_perf,
        total_pn_fy_nsv_payout,

        total_pn_lsv_perf,
        total_pn_fy_lsv_payout,

        None,
        None,

        None,
        None,

        total_pn_total_perf,
        TOTAL_PN_TOTAL_PAYOUT,
    ]

    ws.cell(total_row, 2).value = "Total Target"
    ws.cell(total_row, 2).font = big_bold
    ws.cell(total_row, 2).fill = white_fill
    ws.cell(total_row, 2).border = medium_border
    ws.cell(total_row, 2).alignment = center

    col = 3

    for value in total_pn_values:

        cell = ws.cell(total_row, col)

        cell.value = value

        cell.font = big_bold
        cell.alignment = center
        cell.border = medium_border

        if value is not None:
            cell.number_format = '0.00%'

        if col % 2 == 1:
            cell.fill = pink_fill
        else:
            cell.fill = gray_fill

        col += 1

    # =====================================================
    # BORDERS
    # =====================================================

    for row in range(1, 20):
        for col in range(1, 19):

            ws.cell(row, col).alignment = center

    # =====================================================
    # FREEZE
    # =====================================================

    ws.freeze_panes = "C4"

    for row in range(1, 20):
        ws.cell(row, 2).border = medium_border
        ws.cell(row, 18).border = medium_border

    for col in range(1, 19):
        ws.cell(1, col).border = medium_border
        ws.cell(19, col).border = medium_border